#!/usr/bin/env python3
"""从 VM 拉取深度悬停 CSV 并生成带图表的 xlsx -> D:/Carl_WorkStation/FileData

用法:
  python pull_csv.py            # 一次性拉取全部 CSV 并转换
  python pull_csv.py --watch    # 后台监控模式，每 5 秒自动拉取新文件
  python pull_csv.py --once     # 显式一次性模式（默认行为）
"""
import os, sys, csv, time
import paramiko
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

VM_HOST = '172.16.30.0'
VM_USER = 'carl'
VM_PASS = '159357'
VM_LOG_DIR = '/home/carl/rov_ros2_ws/logs'
LOCAL_DIR = r'D:\Carl_WorkStation\FileData'

# 监控间隔（秒）
WATCH_INTERVAL = 5


def _connect():
    """连接 VM 并返回 (ssh, sftp)"""
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(VM_HOST, username=VM_USER, password=VM_PASS, timeout=10)
    sftp = ssh.open_sftp()
    return ssh, sftp


def csv_to_xlsx(csv_path):
    """将 CSV 转换为带图表的 xlsx"""
    rows = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            rows.append(row)

    if len(rows) < 2:
        return None

    wb = Workbook()

    # ── Sheet 1: 原始数据 ──
    ws = wb.active
    ws.title = '深度数据'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for ci, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            try:
                ws.cell(row=ri, column=ci, value=float(val))
            except (ValueError, TypeError):
                ws.cell(row=ri, column=ci, value=val)

    for ci in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    last_row = len(rows) + 1
    target_depth = rows[0][1]

    # ── Sheet 2: 深度悬停曲线 ──
    ws2 = wb.create_sheet('深度悬停曲线')

    chart1 = LineChart()
    chart1.title = f'深度悬停记录 (目标={target_depth}m)'
    chart1.y_axis.title = '深度 (m)'
    chart1.x_axis.title = '时间 (秒)'
    chart1.style = 10
    chart1.width = 28
    chart1.height = 16

    col_target = 2
    data_target = Reference(ws, min_col=col_target, min_row=1, max_row=last_row)
    chart1.add_data(data_target, titles_from_data=True)

    col_actual = 3
    data_actual = Reference(ws, min_col=col_actual, min_row=1, max_row=last_row)
    chart1.add_data(data_actual, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart1.set_categories(cats)

    chart1.series[0].graphicalProperties.line.solidFill = 'FF0000'
    chart1.series[0].graphicalProperties.line.dashStyle = 'dash'
    chart1.series[1].graphicalProperties.line.solidFill = '0070C0'
    chart1.y_axis.scaling.orientation = 'maxMin'

    ws2.add_chart(chart1, 'A1')

    # 误差统计
    errors = []
    for row in rows:
        try:
            t = float(row[1])
            a = float(row[2])
            errors.append(abs(t - a))
        except (ValueError, IndexError):
            pass
    if errors:
        avg_err = sum(errors) / len(errors)
        max_err = max(errors)
        ws2.cell(row=22, column=1, value='误差统计').font = Font(bold=True, size=12)
        ws2.cell(row=23, column=1, value='平均误差').font = Font(bold=True)
        ws2.cell(row=23, column=2, value=f'{avg_err:.3f} m')
        ws2.cell(row=24, column=1, value='最大误差').font = Font(bold=True)
        ws2.cell(row=24, column=2, value=f'{max_err:.3f} m')
        ws2.cell(row=25, column=1, value='数据点数').font = Font(bold=True)
        ws2.cell(row=25, column=2, value=len(errors))
        ws2.cell(row=26, column=1, value='目标深度').font = Font(bold=True)
        ws2.cell(row=26, column=2, value=f'{target_depth} m')

    # ── Sheet 3: 电机转速 ──
    ws3 = wb.create_sheet('电机转速')
    chart2 = LineChart()
    chart2.title = '电机 RPM 变化'
    chart2.y_axis.title = 'RPM'
    chart2.x_axis.title = '时间 (秒)'
    chart2.style = 10
    chart2.width = 28
    chart2.height = 16

    motor_cols = [6, 7, 8, 9, 10, 11, 12]
    colors = ['1F4E79', '2E75B6', '5B9BD5', '9DC3E6', 'FF6B35', 'FF9B35', '70AD47']

    for mi, (col_idx, color) in enumerate(zip(motor_cols, colors)):
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=last_row)
        chart2.add_data(data, titles_from_data=True)
        chart2.series[mi].graphicalProperties.line.solidFill = color

    chart2.set_categories(cats)
    ws3.add_chart(chart2, 'A1')

    xlsx_path = csv_path.replace('.csv', '.xlsx')
    wb.save(xlsx_path)
    return xlsx_path


def pull_loop(pull_all=False):
    """拉取并转换 CSV。返回新转换的 xlsx 数量。
    
    pull_all=True: 处理所有 CSV（首次/一次性模式）
    pull_all=False: 只处理尚未转换的 CSV（监控模式增量）
    """
    os.makedirs(LOCAL_DIR, exist_ok=True)

    ssh, sftp = _connect()
    try:
        files = sftp.listdir(VM_LOG_DIR)
    except IOError:
        return 0, 0

    csv_files = sorted([f for f in files if f.endswith('.csv')])

    if not csv_files:
        sftp.close()
        ssh.close()
        return 0, 0

    # 拉取所有 CSV
    pulled = 0
    for fname in csv_files:
        remote = f'{VM_LOG_DIR}/{fname}'
        local = os.path.join(LOCAL_DIR, fname)
        try:
            sftp.get(remote, local)
        except Exception:
            continue
        pulled += 1

    sftp.close()
    ssh.close()

    # 转换：全部模式 vs 增量模式
    converted = 0
    for fname in csv_files:
        csv_path = os.path.join(LOCAL_DIR, fname)
        xlsx_path = csv_path.replace('.csv', '.xlsx')

        if not pull_all and os.path.exists(xlsx_path):
            continue  # 监控模式：跳过已转换的

        if os.path.exists(csv_path):
            result = csv_to_xlsx(csv_path)
            if result:
                converted += 1
                print(f'  [{time.strftime("%H:%M:%S")}] 新生成: {os.path.basename(xlsx_path)}')

    return pulled, converted


def main():
    watch_mode = '--watch' in sys.argv
    pull_all = not watch_mode  # 一次性模式拉全部

    if watch_mode:
        print(f'=== 深度悬停 CSV 自动拉取（监控模式） ===')
        print(f'监控间隔: {WATCH_INTERVAL}秒  目标: {LOCAL_DIR}')
        print(f'按 Ctrl+C 停止\n')

        # 首次拉取全部
        prev_csv_count = 0
        pulled, converted = pull_loop(pull_all=True)
        if pulled > 0:
            print(f'  初始拉取: {pulled} CSV, {converted} 新 xlsx')

        try:
            while True:
                time.sleep(WATCH_INTERVAL)
                try:
                    pulled, converted = pull_loop(pull_all=False)
                    if converted > 0:
                        print(f'  [{time.strftime("%H:%M:%S")}] 本轮生成 {converted} 个新 xlsx, 共 {pulled} CSV')
                except Exception as e:
                    print(f'  [{time.strftime("%H:%M:%S")}] 连接失败: {e}，将在 {WATCH_INTERVAL}s 后重试...')
        except KeyboardInterrupt:
            print('\n监控已停止')
    else:
        print(f'=== 深度悬停 CSV → xlsx 一次性拉取 ===')
        pulled, converted = pull_loop(pull_all=True)
        if pulled > 0:
            print(f'\n完成: {pulled} CSV -> {LOCAL_DIR}  ({converted} 个新 xlsx)')
        else:
            print('VM 上无 CSV 文件')


if __name__ == '__main__':
    main()
